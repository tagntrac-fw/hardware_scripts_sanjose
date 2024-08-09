import openpyxl,glob, os, pyshark
from openpyxl import Workbook

wb = Workbook()
# grab the active worksheet
ws = wb.active

#Search for all matching Wireshark captures in directory to parse
index=2
#os.chdir("C:/Python27")

for searchfile in glob.glob("*.pcapng"):
    print("Parsing:", searchfile)
    cap=pyshark.FileCapture(searchfile,display_filter='btle.advertising_address == C0:04:03:45:32:FC')
    name=["File Name: ",searchfile, "\n"]
    
    ws['A1'] = 'File'
    ws['B1'] = 'Test Temp [C]'
    ws['C1'] = 'Packet Seq Number'
    ws['D1'] = 'Payload [Byte]'
    ws['E1'] = 'Channel'
    ws['F1'] = 'RSSI [dBm]'
    ws['G1'] = 'Adv Address'
    ws['H1'] = 'Temp [Hex]'
    ws['I1'] = 'Temp [Decimal]'
    ws['J1'] = 'Tamper Event'
    
    #Single packet per temp
    #x=0
    #for packet in (cap):
    #    if packet.captured_length == '57':
    #        avg_cap=[cap[x]]
    #    x+=1
    
    #5 Packet per temp
    #avg_cap=[cap[0], cap[1], cap[2], cap[3], cap[4], cap[5]]
    
    avg_cap=cap
    
    
    for packet in (avg_cap):
        #print(packet)
        ws.cell(row=index, column=1).value=searchfile
        ws.cell(row=index, column=2).value=int(searchfile[0:2])
        ws.cell(row=index, column=3).value=int(packet.nordic_ble.packet_counter)
        ws.cell(row=index, column=4).value=int(packet.captured_length)
        ws.cell(row=index, column=5).value=int(packet.nordic_ble.channel)
        ws.cell(row=index, column=6).value=int(packet.nordic_ble.rssi)
        ws.cell(row=index, column=7).value=str(packet.btle.advertising_address)
        #print (packet.btle.btcommon_eir_ad_entry_data[6:8])
        if packet.captured_length == '56':
            list=[packet.btle.btcommon_eir_ad_entry_data[9:11],packet.btle.btcommon_eir_ad_entry_data[12:14]]
            s=""
            s=s.join(list)
            ws.cell(row=index, column=8).value=s                #Byte for Temp
            
            #print(int(s,16))
            if int(s,16) < 32767:   #16 bit max positive value
                ws.cell(row=index, column=9).value=0.003906*int(s,16)     
            else:
                ws.cell(row=index, column=9).value=0.003906*(int(s,16)-4096)
            ws.cell(row=index, column=10).value=int(packet.btle.btcommon_eir_ad_entry_data[6:8])        #Tamper Event
        else:
            ws.cell(row=index, column=8).value==''
            ws.cell(row=index, column=9).value==''
            ws.cell(row=index, column=10).value==''
        print(str(packet.btle.btcommon_eir_ad_entry_data.split(':')[11])
              +str(packet.btle.btcommon_eir_ad_entry_data.split(':')[12]))
        index+=1
    #End For
#End For

# Save the file
wb.save("Results.xlsx")
